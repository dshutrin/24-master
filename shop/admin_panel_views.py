from django.shortcuts import render
from django.http import HttpResponseNotFound, HttpResponseRedirect, JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from .forms import *
from .models import *
from django.views.decorators.csrf import csrf_exempt
from django.conf import settings
import os
import uuid
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader


@login_required
def admin_panel(request):
    if request.user.role.access_to_admin_panel:

        return render(request, 'admin_panel/home.html')

    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_roles(request):
    if request.user.role.access_to_admin_panel:

        roles = Role.objects.all()

        return render(request, 'admin_panel/roles.html', {
            'roles': roles
        })

    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_product_categories(request):
    if request.user.role.access_to_admin_panel:

        categories = Category.objects.all()

        return render(request, 'admin_panel/categories.html', {
            'categories': categories
        })

    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_tags(request):
    if request.user.role.access_to_admin_panel:

        tags = Tag.objects.all()

        return render(request, 'admin_panel/tags.html', {
            'tags': tags
        })

    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_products(request):
    if request.user.role.access_to_admin_panel:

        products = Product.objects.all()

        return render(request, 'admin_panel/products.html', {
            'products': products
        })

    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_products_add(request):
    if request.user.role.access_to_admin_panel:
        form = ProductForm()
        if request.method == 'GET':
            return render(request, 'admin_panel/product_add.html', {
                'form': form
            })
        else:
            form = ProductForm(request.POST, request.FILES)
            if form.is_valid:
                form.save()
                return HttpResponseRedirect('/admin_panel/products')
    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_add_tag(request):
    if request.user.role.access_to_admin_panel:
        form = TagForm()
        if request.method == 'GET':
            return render(request, 'admin_panel/tag_add.html', {
                'form': form
            })
        else:
            form = TagForm(request.POST)
            if form.is_valid:
                form.save()
                return HttpResponseRedirect('/admin_panel/tags')
    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_add_category(request):
    if request.user.role.access_to_admin_panel:
        form = CategoryForm()
        if request.method == 'GET':
            return render(request, 'admin_panel/category_add.html', {
                'form': form
            })
        else:
            form = CategoryForm(request.POST)
            if form.is_valid:
                cat = form.save(commit=False)
                cat.is_main = False
                cat.save()
                return HttpResponseRedirect('/admin_panel/categories')
    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_edit_product(request, pid):
    if request.user.role.access_to_admin_panel:

        product = Product.objects.get(id=pid)
        form = ProductForm(instance=product)

        if request.method == 'GET':

            p_tags = [x.tag for x in ProductTag.objects.filter(product=product)]
            ptids = [x.id for x in p_tags]
            all_tags = [x for x in Tag.objects.all().order_by('-id') if x.id not in ptids]

            p_cats = [x.category for x in ProductCategory.objects.filter(product=product)]
            pcids = [x.id for x in p_cats]
            all_cats = [x for x in Category.objects.all().order_by('-id') if x.id not in pcids]

            if product.photo:
                photo_url = product.photo.url
            else:
                photo_url = False

            return render(request, 'admin_panel/product_edit.html', {
                'form': form,
                'photo_url': photo_url,
                'product_tags': p_tags,
                'product_categories': p_cats,
                'all_tags': all_tags,
                'all_categories': all_cats,
                'pid': product.id
            })
        elif request.method == 'POST':
            form = ProductForm(request.POST, request.FILES, instance=product)

            if form.is_valid:
                form.save()
                return HttpResponseRedirect('/admin_panel/products')
    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_panel_remove_product_tag(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            product = Product.objects.get(id=int(request.POST.get('pid')))
            tag = Tag.objects.get(id=int(request.POST.get('tid')))

            ProductTag.objects.filter(product=product, tag=tag).delete()

            return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_panel_add_product_tag(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            product = Product.objects.get(id=int(request.POST.get('pid')))
            tag = Tag.objects.get(id=int(request.POST.get('tid')))

            if ProductTag.objects.filter(product=product, tag=tag).count() == 0:
                ProductTag.objects.create(product=product, tag=tag)

                return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_panel_remove_product_cat(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            product = Product.objects.get(id=int(request.POST.get('pid')))
            cat = Category.objects.get(id=int(request.POST.get('cid')))

            ProductCategory.objects.filter(product=product, category=cat).delete()

            return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_panel_add_product_cat(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            product = Product.objects.get(id=int(request.POST.get('pid')))
            cat = Category.objects.get(id=int(request.POST.get('cid')))

            if ProductCategory.objects.filter(product=product, category=cat).count() == 0:
                ProductCategory.objects.create(product=product, category=cat)

                return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def delete_product(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':
            try:
                Product.objects.filter(id=int(request.POST.get('pid'))).delete()
                return JsonResponse({}, status=200)
            except:
                return JsonResponse({}, status=224)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_panel_delete_category(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            cid = int(request.POST.get('cid'))

            category = Category.objects.filter(id=cid)
            cats = CatImage.objects.filter(category=category[0])
            if len(cats) > 0:
                if cats[0].image:
                    if os.path.exists(cats[0].image.url):
                        os.remove(cats[0].image.url)
                cats[0].delete()
            category.delete()
            return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_panel_delete_tag(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            cid = int(request.POST.get('tid'))

            Tag.objects.filter(id=cid).delete()
            return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def delete_order(request):
    if request.user.role.access_to_admin_panel:
        if request.method == 'POST':

            Order.objects.filter(id=int(request.POST.get('oid'))).delete()
            return JsonResponse({}, status=200)

    else:
        return HttpResponseNotFound(request)


@login_required
def admin_panel_orders(request):
    if request.user.role.access_to_admin_panel:
        orders = Order.objects.all().order_by('-create_at')
        return render(request, 'admin_panel/orders.html', {'orders': orders})


@login_required
def admin_panel_order(request, oid):
    if request.user.role.access_to_admin_panel:
        order = Order.objects.get(id=oid)
        order_products = OrderElement.objects.filter(order=order)
        for prod in order_products:
            prod.total_price = prod.product.price * prod.pcount
            print(prod.product.price, prod.pcount, prod.total_price, order.total_price)
        return render(request, 'admin_panel/order_detail.html', {'order': order, 'order_products': order_products})
    else:
        return HttpResponseNotFound(request)


@csrf_exempt
@login_required
def admin_delete_role(request, rid):
    if request.user.role.access_to_admin_panel:
        role = Role.objects.get(id=rid)
        role.delete()
        return HttpResponse(status=200)
    else:
        return HttpResponseNotFound(request)


@login_required
def admin_edit_role(request, rid):
    if request.user.role.access_to_admin_panel:

        role = Role.objects.get(id=rid)
        form = RoleForm(instance=role)

        if request.method == 'GET':

            return render(request, 'admin_panel/edit_role.html', {
                'form': form
            })
        elif request.method == 'POST':
            form = RoleForm(request.POST, instance=role)

            if form.is_valid:
                form.save()
                return HttpResponseRedirect('/admin_panel/roles')
    else:
        return HttpResponseNotFound(request)


@login_required
def admin_add_role(request):
    if request.user.role.access_to_admin_panel:

        form = RoleForm()

        if request.method == 'GET':

            return render(request, 'admin_panel/add_role.html', {
                'form': form
            })
        elif request.method == 'POST':
            form = RoleForm(request.POST)

            if form.is_valid:
                form.save()
                return HttpResponseRedirect('/admin_panel/roles')
    else:
        return HttpResponseNotFound(request)


@login_required
def admin_edit_order(request, id):
    if request.user.role.access_to_admin_panel:

        order = Order.objects.get(id=id)
        form = OrderForm(instance=order)
        total_price = 0

        if request.method == 'GET':
            order_products = OrderElement.objects.filter(order=order)
            for prod in order_products:
                prod.total_price = prod.product.price * prod.pcount
                prod.save()
                total_price += prod.total_price


            return render(request, 'admin_panel/order_edit.html', {
                'form': form,
                'order_products': order_products,
                'total_price': total_price
            })
        elif request.method == 'POST':
            form = OrderForm(request.POST, instance=order)


            if form.is_valid:
                form.save()
                return HttpResponseRedirect('/admin_panel/orders')
    else:
        return HttpResponseNotFound(request)


@csrf_exempt
def order_upcount(request):
    if request.POST:
        id = int(request.POST.get('id'))
        order_el = OrderElement.objects.get(id=id)
        order_el.pcount += 1
        order = order_el.order
        order.total_price += order_el.product.price
        order.save()
        order_el.save()

        return JsonResponse({'price': order_el.product.price}, status=200)


@csrf_exempt
def order_downcount(request):
    if request.POST:
        id = int(request.POST.get('id'))
        order_el = OrderElement.objects.get(id=id)
        if order_el.pcount > 0:
            order_el.pcount -= 1
            order = order_el.order
            order.total_price -= order_el.product.price
            order.save()
            order_el.save()
            return JsonResponse({'price': order_el.product.price}, status=200)
        else:
            return HttpResponse(status=203)


@csrf_exempt
def cat_do_main(request):
    if request.POST:
        cid = int(request.POST.get('cid'))
        if len(Category.objects.filter(show_on_home=True)) < 5:
            category = Category.objects.get(id=cid)
            category.show_on_home = True
            category.save()
            return JsonResponse({}, status=200)
        else:
            return JsonResponse({}, status=220)


@csrf_exempt
def cat_do_common(request):
    if request.POST:
        cid = int(request.POST.get('cid'))
        category = Category.objects.get(id=cid)
        cats = CatImage.objects.filter(category=category)
        if len(cats) > 0:
            if cats[0].image:
                if os.path.exists(cats[0].image.url):
                    os.remove(cats[0].image.url)
            cats[0].delete()
        category.show_on_home = False
        category.save()
        return JsonResponse({}, status=200)


def get_ages_from_age(age):
    age_start = None
    age_end = None
    if age:
        integers = []
        n = 0
        while n < len(age):
            s_int = ''
            while n < len(age) and '0' <= age[n] <= '9':
                s_int += age[n]
                n += 1
            n += 1
            if s_int != '':
                integers.append(int(s_int))
        age_start = min(integers)
        if len(integers) > 1:
            age_end = max(integers)

    return age_start, age_end


@login_required
@csrf_exempt
def add_price(request):
    errs = []
    if request.method == 'GET':
        return render(request, 'admin_panel/add_price.html', {'errs': errs})

    elif request.method == 'POST':
        f = request.FILES.get("file")

        try:
            with open(os.path.join(settings.BASE_DIR, "media/tmp/prais.xlsx").replace('\\', '/'), "wb+") as destination:
                for chunk in f.chunks():
                    destination.write(chunk)

            book = load_workbook(filename=os.path.join(settings.BASE_DIR, "media/tmp/prais.xlsx").replace('\\', '/'))
            sheet = book['П Р А Й С (ПТО)']
            image_loader = SheetImageLoader(sheet)

            codes = []
            dps = Product.objects.all()

            for i in range(48, book.active.max_row):
                try:
                    i = str(i)
                    product_code = sheet['A' + i].value
                    name = sheet['B' + i].value
                    age = sheet['C' + i].value
                    try:
                        image = image_loader.get('D' + i)
                    except:
                        image = None
                    height = sheet['E' + i].value
                    length = sheet['F' + i].value
                    width = sheet['G' + i].value
                    params = sheet['H' + i].value
                    weight = sheet['I' + i].value
                    concrete = sheet['J' + i].value
                    installation_time = sheet['K' + i].value
                    price = sheet['L' + i].value
                    age_start, age_end = get_ages_from_age(age)

                    codes.append(str(product_code))
                    a = True
                    for prod in dps.filter(product_code=product_code):
                        if prod.product_code == product_code:
                            a = False
                            prod.name = name
                            prod.price = price
                            if height:
                                try:
                                    if str(height).isdigit():
                                        prod.height = height
                                    else:
                                        prod.height = None
                                except:
                                    prod.height = None
                            if length:
                                try:
                                    if str(length).isdigit():
                                        prod.length = length
                                    else:
                                        prod.length = None
                                except:
                                    prod.length = None
                            if width:
                                try:
                                    if str(width).isdigit():
                                        prod.width = width
                                    else:
                                        prod.width = None
                                except:
                                    prod.width = None
                            if weight:
                                try:
                                    prod.weight = round(float(weight), 2)
                                except:
                                    prod.weight = None
                            if params:
                                try:
                                    prod.params = params
                                except:
                                    prod.params = None
                            if concrete:
                                try:
                                    prod.concrete = round(float(concrete), 2)
                                except:
                                    prod.concrete = None
                            if installation_time:
                                try:
                                    prod.installation_time = round(float(weight), 2)
                                except:
                                    prod.weight = None
                            if age_start:
                                try:
                                    prod.age_start = age_start
                                except:
                                    pass
                            if age_end:
                                try:
                                    prod.age_end = age_end
                                except:
                                    pass
                            if image:
                                img_url = 'image_' + str(uuid.uuid1()) + '.png'
                                if prod.photo:
                                    if os.path.exists(os.path.join(str(settings.BASE_DIR) + prod.photo.url).replace('\\', '/')):
                                        os.remove(os.path.join(str(settings.BASE_DIR) + prod.photo.url).replace('\\', '/'))
                                try:
                                    image.save(os.path.join(settings.BASE_DIR, 'media/products_images/' + img_url))
                                    prod.photo = 'products_images/' + img_url
                                except:
                                    prod.photo = None
                            prod.save()
                            break
                    if a:
                        product = Product()
                        product.product_code = product_code
                        product.name = name
                        product.price = price
                        if height:
                            try:
                                if str(height).isdigit():
                                    product.height = height
                            except:
                                pass
                        if length:
                            try:
                                if str(length).isdigit():
                                    product.length = length
                            except:
                                pass
                        if width:
                            try:
                                if str(width).isdigit():
                                    product.width = width
                            except:
                                pass
                        if weight:
                            try:
                                product.weight = round(float(weight), 2)
                            except:
                                pass
                        if params:
                            product.params = str(params)
                        if concrete:
                            try:
                                product.concrete = round(float(concrete), 2)
                            except ValueError:
                                pass
                        if installation_time:
                            try:
                                product.installation_time = round(float(installation_time), 2)
                            except ValueError:
                                pass
                        if age_start:
                            product.age_start = age_start
                        if age_end:
                            product.age_end = age_end
                        if image:
                            img_url = 'image_' + str(uuid.uuid1()) + '.png'
                            try:
                                image.save(os.path.join(settings.BASE_DIR, 'media/products_images/' + img_url))
                                product.photo = f'products_images/{img_url}'
                            except:
                                pass
                        product.save()
                except:
                    errs.append(f'Не получилось сохранить товар из строки {i}')

                if os.path.exists(os.path.join(settings.BASE_DIR, "media/tmp/prais.xlsx").replace('\\', '/')):
                    os.remove(os.path.join(settings.BASE_DIR, "media/tmp/prais.xlsx").replace('\\', '/'))
                del_codes = [x.product_code for x in Product.objects.all() if not (x.product_code in codes)]

                for cod in del_codes:
                    try:
                        Product.objects.get(product_code=cod).delete()
                    except:
                        errs.append(f'Не получилось удалить товар {cod}!')

        except:
            errs.append('Что-то пошло не так!')

        return render(request, 'admin_panel/add_price.html', {'errs': errs})


@csrf_exempt
def set_cat_image(request):
    if request.POST:
        print(str(request.POST.get('formdata')))
        '''cats = Category.objects.filter(id=request.POST.get('cid'), show_on_home=True)
        if len(cats) > 0:
            cat = cats[0]
            img = request.POST['formdata'][0]
            print(img)
            if img:
                CatImage(category=cat, image=img).save()
                return JsonResponse({}, status=200)
            else:
                return JsonResponse({}, status=204)'''

    return JsonResponse({}, status=204)


@login_required
@csrf_exempt
def admin_edit_cat(request, cid):
    if request.user.role.access_to_admin_panel:
        category = Category.objects.filter(id=cid, show_on_home=True)
        if len(category) > 0:
            category = category[0]
            if len(CatImage.objects.filter(category=category)) > 0:
                form = MainCatForm(instance=CatImage.objects.filter(category=category)[0])

                if request.method == 'GET':
                    return render(request, 'admin_panel/cat_edit.html', {
                        'form': form
                    })
                elif request.method == 'POST':
                    form = MainCatForm(request.POST, request.FILES, instance=CatImage.objects.filter(category=category)[0])


                    if form.is_valid:
                        form.save()
                        return HttpResponseRedirect('/admin_panel/categories')
            else:
                form = MainCatForm()

                if request.method == 'GET':
                    return render(request, 'admin_panel/cat_edit.html', {
                        'form': form
                    })
                elif request.method == 'POST':
                    form = MainCatForm(request.POST, request.FILES)


                    if form.is_valid:
                        catim = form.save(commit=False)
                        catim.category = category
                        catim.save()
                        return HttpResponseRedirect('/admin_panel/categories')
        else:
            return HttpResponseNotFound(request)
    else:
        return HttpResponseNotFound(request)
